"""
Excel Report Generator for Mannaz Page Freshness Monitor
Creates professional Excel reports with charts and multiple worksheets
"""

import pandas as pd
from datetime import datetime
import logging

# Try to import Excel libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel reports disabled.")


def create_excel_report(current_df, previous_df, changes_df, output_file):
    """
    Create comprehensive Excel report with multiple worksheets
    
    Args:
        current_df: DataFrame with current page data
        previous_df: DataFrame with previous month data (can be None)
        changes_df: DataFrame with page changes (can be None)
        output_file: Path to save Excel file
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
    
    logger = logging.getLogger(__name__)
    logger.info(f"Creating Excel report: {output_file}")
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Dashboard with summary statistics
        create_dashboard_sheet(writer, current_df, previous_df, changes_df)
        
        # Sheet 2: All current pages
        create_all_pages_sheet(writer, current_df)
        
        # Sheet 3: Status changes (if available)
        if changes_df is not None and not changes_df.empty:
            create_changes_sheet(writer, changes_df)
        
        # Sheet 4: Updated articles (Fresh items)
        create_fresh_pages_sheet(writer, current_df)
        
        # Sheet 5: Pages needing attention (Rotting + Outdated)
        create_attention_sheet(writer, current_df)
    
    logger.info(f"Excel report created: {output_file}")


def create_dashboard_sheet(writer, current_df, previous_df, changes_df):
    """Create dashboard summary sheet"""
    
    # Calculate statistics
    total_pages = len(current_df)
    fresh_count = len(current_df[current_df['freshness'] == 'Fresh'])
    rotting_count = len(current_df[current_df['freshness'] == 'Rotting'])
    outdated_count = len(current_df[current_df['freshness'] == 'Outdated'])
    
    # Create summary DataFrame
    summary_data = {
        'Metric': [
            'Total Pages',
            'Fresh (< 6 months)',
            'Rotting (6-12 months)',
            'Outdated (> 12 months)',
            'Fresh %',
            'Needs Attention %'
        ],
        'Count': [
            total_pages,
            fresh_count,
            rotting_count,
            outdated_count,
            f"{(fresh_count/total_pages*100):.1f}%" if total_pages > 0 else "0%",
            f"{((rotting_count + outdated_count)/total_pages*100):.1f}%" if total_pages > 0 else "0%"
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Dashboard', index=False, startrow=2)
    
    # Access the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Dashboard']
    
    # Add title
    worksheet['A1'] = 'Mannaz Page Freshness Report'
    worksheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    worksheet['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    worksheet.merge_cells('A1:B1')
    
    # Add date
    worksheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    # Format summary table
    for row in range(3, 9):
        worksheet[f'A{row}'].font = Font(bold=True)
        worksheet[f'B{row}'].alignment = Alignment(horizontal='right')
    
    # Add pie chart for freshness distribution
    if total_pages > 0:
        pie = PieChart()
        labels = Reference(worksheet, min_col=1, min_row=4, max_row=6)
        data = Reference(worksheet, min_col=2, min_row=3, max_row=6)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Content Freshness Distribution"
        worksheet.add_chart(pie, "D3")
    
    # Add changes summary if available
    if changes_df is not None and not changes_df.empty:
        new_pages = len(changes_df[changes_df['status_change'] == 'New page'])
        improvements = len(changes_df[changes_df['status_change'].str.contains('✅', na=False)])
        degradations = len(changes_df[changes_df['status_change'].str.contains('⚠️|🚨', na=False, regex=True)])
        
        changes_data = {
            'Change Type': ['New Pages', 'Improvements', 'Degradations'],
            'Count': [new_pages, improvements, degradations]
        }
        changes_summary = pd.DataFrame(changes_data)
        changes_summary.to_excel(writer, sheet_name='Dashboard', index=False, startrow=11)
        
        # Format changes section
        worksheet['A11'] = 'Monthly Changes'
        worksheet['A11'].font = Font(size=12, bold=True)


def create_all_pages_sheet(writer, current_df):
    """Create sheet with all current pages"""
    
    # Select and order columns
    columns = ['title', 'path', 'freshness', 'language', 'date_modified', 'date_created', 'tags']
    export_df = current_df[[col for col in columns if col in current_df.columns]].copy()
    
    # Sort by freshness (Outdated first)
    freshness_order = {'Outdated': 0, 'Rotting': 1, 'Fresh': 2}
    export_df['sort_order'] = export_df['freshness'].map(freshness_order)
    export_df = export_df.sort_values('sort_order').drop('sort_order', axis=1)
    
    # Write to Excel
    export_df.to_excel(writer, sheet_name='All Pages', index=False)
    
    # Format worksheet
    worksheet = writer.sheets['All Pages']
    
    # Format header
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Color code freshness column
    freshness_col = None
    for idx, col in enumerate(worksheet[1], 1):
        if col.value == 'freshness':
            freshness_col = idx
            break
    
    if freshness_col:
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=freshness_col)
            if cell.value == 'Fresh':
                cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            elif cell.value == 'Rotting':
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            elif cell.value == 'Outdated':
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_changes_sheet(writer, changes_df):
    """Create sheet showing status changes"""
    
    # Select relevant columns
    columns = ['title', 'path', 'status_change', 'freshness_previous', 'freshness_current', 'date_modified']
    export_df = changes_df[[col for col in columns if col in changes_df.columns]].copy()
    
    # Write to Excel
    export_df.to_excel(writer, sheet_name='Status Changes', index=False)
    
    # Format worksheet
    worksheet = writer.sheets['Status Changes']
    
    # Format header
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_fresh_pages_sheet(writer, current_df):
    """Create sheet with recently updated pages"""
    
    fresh_df = current_df[current_df['freshness'] == 'Fresh'].copy()
    
    if fresh_df.empty:
        # Create empty sheet with message
        pd.DataFrame({'Message': ['No fresh pages found']}).to_excel(
            writer, sheet_name='Updated Articles', index=False
        )
        return
    
    # Sort by modification date (newest first)
    if 'date_modified' in fresh_df.columns:
        fresh_df = fresh_df.sort_values('date_modified', ascending=False)
    
    # Select columns
    columns = ['title', 'path', 'date_modified', 'tags']
    export_df = fresh_df[[col for col in columns if col in fresh_df.columns]]
    
    export_df.to_excel(writer, sheet_name='Updated Articles', index=False)
    
    # Format worksheet
    worksheet = writer.sheets['Updated Articles']
    
    # Format header
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_attention_sheet(writer, current_df):
    """Create sheet with pages needing attention"""
    
    attention_df = current_df[
        (current_df['freshness'] == 'Rotting') | 
        (current_df['freshness'] == 'Outdated')
    ].copy()
    
    if attention_df.empty:
        pd.DataFrame({'Message': ['All pages are fresh!']}).to_excel(
            writer, sheet_name='Needs Attention', index=False
        )
        return
    
    # Sort by freshness and date (Outdated first, oldest first)
    freshness_order = {'Outdated': 0, 'Rotting': 1}
    attention_df['sort_order'] = attention_df['freshness'].map(freshness_order)
    
    if 'date_modified' in attention_df.columns:
        attention_df = attention_df.sort_values(['sort_order', 'date_modified'])
    else:
        attention_df = attention_df.sort_values('sort_order')
    
    attention_df = attention_df.drop('sort_order', axis=1)
    
    # Select columns
    columns = ['title', 'path', 'freshness', 'date_modified', 'tags']
    export_df = attention_df[[col for col in columns if col in attention_df.columns]]
    
    export_df.to_excel(writer, sheet_name='Needs Attention', index=False)
    
    # Format worksheet
    worksheet = writer.sheets['Needs Attention']
    
    # Format header
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Color code by freshness
    freshness_col = None
    for idx, col in enumerate(worksheet[1], 1):
        if col.value == 'freshness':
            freshness_col = idx
            break
    
    if freshness_col:
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=freshness_col)
            if cell.value == 'Rotting':
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            elif cell.value == 'Outdated':
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    print("This module should not be run directly.")
    print("Import it in your main script: from generate_excel_report import create_excel_report")
